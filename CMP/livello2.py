from PyQt6.QtCore import Qt, QFile, QTextStream, QSize, pyqtSignal, pyqtSlot, QTimer
from PyQt6.QtWidgets import (
    QPushButton, QLabel, QSizePolicy, QVBoxLayout, QWidget, QScrollArea, QFrame, QHBoxLayout, QDialog, QGridLayout, QStackedWidget, QMenuBar, 
    QSpinBox, QRadioButton, QFileDialog, QCheckBox
)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import csv
from datetime import datetime
import threading
from PyQt6.QtGui import QPixmap, QAction, QIcon, QColor
import os 
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas


from PAGE import setting_page as s
from API import funzioni as f, LOG as l 
from CMP import puls_livello2 as p, puls_livello1_interni as pi, messaggi as m
from OBJ import thread_pesata_continua as t 


class Livello2(QWidget):
    
    dir_path = f.get_resource_path(os.path.join("data", "csv"))
    
    
    def __init__(self, master):
        super().__init__()
        
        self.master:s.Settings = master
        # Per memorizzare i dati della pesata
        self.pesi_totali = {}
        self.pesi_celle = {}
        
        self.main_layout = QVBoxLayout()
        self.stacked_widget = QStackedWidget()
        
        self.home_page()        #p0
        self.diagnosi_page()    #p1
        self.db_page()          #p2
        self.log_page()         #p3
        self.sett_page()        #p4
        
        self.main_layout.addWidget(self.stacked_widget)
        self.setLayout(self.main_layout)
        self.set_background_color()
        
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.start_pesata)
        self.stop_requested = False
        
        #self.setMaximumSize(self.master.master.screen_width,(self.master.master.screen_height))  # Imposta una dimensione fissa per la finestra principale
        # self.stacked_widget.setSizePolicy(QSizePolicy.Policy.Fixed, QSizePolicy.Policy.Fixed)
        
    def set_background_color(self):
        p = self.palette()
        p.setColor(self.backgroundRole(), QColor.fromRgb(254,254,254))
        self.setPalette(p)
        self.load_stylesheet()

    def load_stylesheet(self):
        file = QFile(f.get_style("settings.qss"))
        if file.open(QFile.OpenModeFlag.ReadOnly | QFile.OpenModeFlag.Text):
            stream = QTextStream(file)
            style_sheet = stream.readAll()
            file.close()
            self.setStyleSheet(style_sheet)
    
    def home_page(self):
        home_widget = QWidget()
        home_layout = QVBoxLayout()
        home_widget.setLayout(home_layout)
        
        p1 = p.ClickableWidget(f.get_img("DIAGNOSI.png"), "Misurazione Continua")
        p1.clicked.connect(self.show_diagnosi_page)
        
        p2 = p.ClickableWidget(f.get_img("PARAM_SET.png"), "Settaggio Parametri")
        p2.clicked.connect(self.show_param_page)
        
        
        p3 = p.ClickableWidget(f.get_img("DB_F.png"), "Fuzioni Database")
        p3.clicked.connect(self.db_operatione)
        
        p4 = p.ClickableWidget(f.get_img("LOG_DATA.png"), "Funzioni DataLog")
        p4.clicked.connect(self.dataLog_operatione)
        
        p5 = p.ClickableWidget(f.get_img("MOD_IMPO.png"), "Funzioni Mod-Bus")
        
        home_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        home_layout.addStretch()
        home_layout.addWidget(p1)
        home_layout.addSpacing(10)
        home_layout.addWidget(p2)
        home_layout.addSpacing(10)
        home_layout.addWidget(p3)
        home_layout.addSpacing(10)
        home_layout.addWidget(p4)
        home_layout.addSpacing(10)
        home_layout.addWidget(p5)
        home_layout.addStretch()
        
        self.stacked_widget.addWidget(home_widget)
        
    def diagnosi_page(self):
        diagnosi_widget = QWidget()
        diagnosi_layout = QVBoxLayout()
        diagnosi_widget.setLayout(diagnosi_layout)
        
        h = QHBoxLayout()
        lab = QLabel("IMPOSTAZIONI DEL TEST : ")
        lab.setObjectName("titoloi")
        
        self.tempo = QSpinBox(self)
        self.tempo.setRange(1, 15)
        self.tempo.setValue(1)
        self.tempo.setSingleStep(1)
        self.tempo.setMinimumWidth(100)
        
        tempolab = QLabel("minuti")
        tempolab.setObjectName("desc1")
        
        self.csv = QRadioButton("CSV")
        self.csv.setChecked(True)
        
        h.addWidget(lab)
        h.addStretch()
        h.addWidget(self.tempo)
        h.addWidget(tempolab)
        h.addStretch()
        h.addWidget(self.csv)
        h.addStretch()
        
        h2 = QHBoxLayout()
        
        self.start = QPushButton("START")
        self.start.setObjectName("big")
        self.start.clicked.connect(self.start_reg)
        self.stop = QPushButton("STOP")
        self.stop.setObjectName("bigd")
        self.stop.clicked.connect(self.stop_reg)
        self.stampa = QPushButton("STAMPA")
        self.stampa.setObjectName("big")
        self.stampa.setVisible(False)
        self.stampa.clicked.connect(self.stampa_reg)
        
        labmi = QLabel("MISURE EFFETTUATE:")
        labmi.setObjectName("desc1")
        
        self.mi = QLabel("--")
        self.mi.setObjectName("pass")
        self.mi.setMinimumWidth(60)
        
        h2.addWidget(self.start)
        h2.addWidget(self.stop)
        h2.addWidget(self.stampa)
        h2.addStretch()
        h2.addWidget(labmi)
        h2.addWidget(self.mi)

        diagnosi_layout.addLayout(h)
        diagnosi_layout.addLayout(h2)
        diagnosi_layout.addSpacing(20)
        
        # Create a QScrollArea to contain the graphs
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        
        # Create a QWidget to hold the layout
        graph_container = QWidget()
        self.graph_layout = QVBoxLayout(graph_container)
        
        # Set the graph_container widget as the scroll area's widget
        scroll_area.setWidget(graph_container)
        
        # Add the scroll area to the diagnosi layout
        diagnosi_layout.addWidget(scroll_area)
        
        
        self.stacked_widget.addWidget(diagnosi_widget)
        
    def start_reg(self):
        print("start")
        self.stop.setObjectName("big")
        self.start.setObjectName("bigd")
        self.stampa.setVisible(False)
        self.start.style().unpolish(self.start)
        self.start.style().polish(self.start)
        self.start.update()

        self.stop.style().unpolish(self.stop)
        self.stop.style().polish(self.stop)
        self.stop.update()

        self.start.setEnabled(False)
        self.stop.setEnabled(True)
        
        self.stop_requested = False
        self.start_pesata()
        

    def stop_reg(self):
        print("stop")
        self.stop_requested = True
        self.timer.stop()
        self.stop.setObjectName("bigd")
        self.start.setObjectName("big")

        self.start.style().unpolish(self.start)
        self.start.style().polish(self.start)
        self.start.update()

        self.stop.style().unpolish(self.stop)
        self.stop.style().polish(self.stop)
        self.stop.update()
        self.stampa.setVisible(True)
        self.start.setEnabled(True)
        self.stop.setEnabled(False)
        
    def show_diagnosi_page(self):
        self.stacked_widget.setCurrentIndex(1)
        self.master.contro_label(3)
        
    def show_param_page(self):
        self.stacked_widget.setCurrentIndex(4)
        self.master.contro_label(8)
        
    def home(self):
        self.stacked_widget.setCurrentIndex(0)
        self.master.contro_label(2)
        
        
    def start_pesata(self):
        # Avvia il thread
        if len(self.master.master.lista_bilance) != 0 and not self.stop_requested:
            self.pesata_thread = t.PesataThread(self.master.master)
            self.pesata_thread.pesata_completata.connect(self.on_pesata_completata)
            self.pesata_thread.start()

    @pyqtSlot(list)
    def on_pesata_completata(self, pesi_bilance):
        self._log_thread_info("on_pesata_completata")

        # Ferma il thread e la barra di progresso
        self.pesata_thread.quit()

        print(f'DEBUG LIVELLO2 | {pesi_bilance}')

        for n, pesata in enumerate(pesi_bilance):
            bilancia_dict = self.convert_to_dict(pesata, n)
            self.create_csv_file()
            self.write_to_csv(bilancia_dict)

            # Aggiorna i dati per i grafici
            if n not in self.pesi_totali:
                self.pesi_totali[n] = []
                self.pesi_celle[n] = {1: [], 2: [], 3: [], 4: []}

            self.pesi_totali[n].append(bilancia_dict['pesoTot'])
            for i in range(1, 5):
                if bilancia_dict[f'pesoc{i}'] != '':
                    self.pesi_celle[n][i].append(bilancia_dict[f'pesoc{i}'])

            # Aggiungi o aggiorna i grafici
            self.update_graph(n)

        if not self.stop_requested:
            self.timer.start(self.tempo.value() * 60 * 1000)  # Convertire i minuti in millisecondi

    def update_graph(self, bilancia_id):
        """Aggiorna i grafici per una specifica bilancia."""
    
        # Define colors for the four cells
        colors = ['r', 'g', 'b', 'm']  # Red, Green, Blue, Magenta (or choose any colors you prefer)
    
        # Crea una figura e due subplot per il peso totale e i pesi delle celle
        fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(5, 8))
        fig.suptitle(f"Bilancia {bilancia_id + 1}")
    
        # Grafico peso totale
        ax1.plot(self.pesi_totali[bilancia_id], label='Peso Totale', marker='o')
        ax1.set_ylabel("Peso Totale")
        ax1.legend()
    
        # Grafico pesi celle
        for i in range(1, 5):
            if len(self.pesi_celle[bilancia_id][i]) > 0:
                ax2.plot(self.pesi_celle[bilancia_id][i], label=f'Cella {i}', color=colors[i-1], marker='o')
    
        ax2.set_ylabel("Pesi Celle")
        ax2.legend()
    
        # If a previous graph exists, update it
        if self.graph_layout.count() > bilancia_id:
            graph_container = self.graph_layout.itemAt(bilancia_id).widget()
            canvas = graph_container.layout().itemAt(1).widget()
            canvas.figure.clf()
            ax1, ax2 = canvas.figure.add_subplot(2, 1, 1), canvas.figure.add_subplot(2, 1, 2)
            ax1.plot(self.pesi_totali[bilancia_id], label='Peso Totale', marker='o')
            ax1.set_ylabel("Peso Totale")
            ax1.legend()
            for i in range(1, 5):
                if len(self.pesi_celle[bilancia_id][i]) > 0:
                    ax2.plot(self.pesi_celle[bilancia_id][i], label=f'Cella {i}', color=colors[i-1], marker='o')
            ax2.set_ylabel("Pesi Celle")
            ax2.legend()
            canvas.draw()
        else:
            # Otherwise, create a new graph
            canvas = FigureCanvas(fig)
            container = QVBoxLayout()
            label = QLabel(f"Bilancia {bilancia_id + 1}")
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            container.addWidget(label)
            container.addWidget(canvas)
            canvas.setMinimumHeight(250)
            graph_widget = QWidget()
            graph_widget.setLayout(container)
            self.graph_layout.addWidget(graph_widget)
    
        canvas.draw()
        
        try:
            current_value = int(self.mi.text())
        except ValueError:
            # If not, set it to 0
            current_value = 0

        # Increment and update the QLabel
        self.mi.setText(str(current_value + 1))
        
            
            
        
    def _log_thread_info(self, function_name):
        """Log thread information for diagnostics."""
        current_thread = threading.current_thread()
        print(f"DEBUG THREAD | {function_name} eseguito su thread: {current_thread.name} (ID: {current_thread.ident})")
        
    def create_csv_file(self):
        """Crea un file CSV con la data odierna nel nome all'interno della cartella specificata."""
        today = datetime.now().strftime("%d%m%y")
        self.csv_file_path = os.path.join(self.dir_path, f"TEST{today}.csv")

        # Verifica se il file esiste già
        if not os.path.isfile(self.csv_file_path):
            with open(self.csv_file_path, mode='w', newline='') as file:
                writer = csv.writer(file)
                # Scrive l'intestazione del file
                writer.writerow(["id_bilancia", "pesoTot", "pesoc1", "pesoc2", "pesoc3", "pesoc4", "ora"])
            print(f"DEBUG | File creato: {self.csv_file_path}")
        else:
            print(f"DEBUG | Il file esiste già: {self.csv_file_path}")

    def write_to_csv(self, bilancia_dict):
        """Scrive i dati della pesata nel file CSV."""
        with open(self.csv_file_path, mode='a', newline='') as file:
            writer = csv.writer(file)
            current_time = datetime.now().strftime("%H:%M:%S")
            row = [
                bilancia_dict['id_bilancia'],
                bilancia_dict['pesoTot'],
                bilancia_dict.get('pesoc1', ''),
                bilancia_dict.get('pesoc2', ''),
                bilancia_dict.get('pesoc3', ''),
                bilancia_dict.get('pesoc4', ''),
                current_time
            ]
            writer.writerow(row)
            print(f"DEBUG | Dati scritti nel file: {row}")
            
    def stampa_reg(self):
        massimo = []
        minimo = []
        media_pesi_totali = []
        media_pesi_celle = [[] for _ in range(4)]  # List for each cell (1-4)

        for n in range(len(self.pesi_totali)):
            # Ensure the lists are not empty before performing operations
            if self.pesi_totali[n]:  
                massimo.append(max(self.pesi_totali[n]))
                minimo.append(min(self.pesi_totali[n]))
                media_pesi_totali.append(sum(self.pesi_totali[n]) / len(self.pesi_totali[n]))
            else:
                massimo.append(None)
                minimo.append(None)
                media_pesi_totali.append(None)

            for i in range(1, 5):  # Loop over each cell (1-4)
                if self.pesi_celle[n][i]:  # Check if the list is not empty
                    cell_max = max(self.pesi_celle[n][i])
                    cell_min = min(self.pesi_celle[n][i])
                    cell_media = sum(self.pesi_celle[n][i]) / len(self.pesi_celle[n][i])
                else:
                    cell_max = None
                    cell_min = None
                    cell_media = None

                media_pesi_celle[i-1].append((cell_max, cell_min, cell_media))

        self.write_to_excel(massimo, minimo, media_pesi_totali, media_pesi_celle)
    
    def write_to_excel(self, massimo, minimo, media_pesi_totali, media_pesi_celle):
        # Create a new Excel workbook and sheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Pesi Bilance"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        subtitle_font = Font(bold=True, color="000000")
        subtitle_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        max_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # Light red for massimo
        min_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # Light green for minimo
        media_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")  # Light blue for media

        # Write data for each balance
        balance_count = len(self.pesi_totali)
        row = 1
        for n in range(balance_count):
            # Header for each balance
            ws.cell(row=row, column=1, value=f"BILANCIA {n+1}")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            for cell in ws[row]:
                cell.font = header_font
                cell.fill = header_fill
            row += 1

            # Number of weights recorded
            num_pesi = len(self.pesi_totali[n])
            ws.cell(row=row, column=1, value=f"NUMERO PESI REGISTRATI: {num_pesi} tempo: {num_pesi * self.tempo.value()} min")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
            for cell in ws[row]:
                cell.font = subtitle_font
                cell.fill = subtitle_fill
            row += 1

            # Total weight data
            ws.cell(row=row, column=1, value="PESO TOTALE")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1

            # Media row
            ws.cell(row=row, column=1, value="MEDIA:")
            ws.cell(row=row, column=3, value=f"{media_pesi_totali[n] if media_pesi_totali[n] is not None else 'N/A'}")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.cell(row=row, column=1).fill = media_fill  # Apply fill color to media description cell
            row += 1

            # Max-Min row
            ws.cell(row=row, column=1, value="MAX-MIN:")
            ws.cell(row=row, column=2, value="massimo:")
            ws.cell(row=row, column=3, value=f"{massimo[n] if massimo[n] is not None else 'N/A'}")
            ws.cell(row=row, column=4, value="minimo:")
            ws.cell(row=row, column=5, value=f"{minimo[n] if minimo[n] is not None else 'N/A'}")
            ws.cell(row=row, column=2).fill = max_fill  # Apply fill color to massimo description cell
            ws.cell(row=row, column=4).fill = min_fill  # Apply fill color to minimo description cell
            row += 1

            # Data for each cell
            for i in range(4):
                max_cella, min_cella, media_cella = media_pesi_celle[i][n]

                # Row 1: "CELLA INTERNA {i+1}"
                ws.cell(row=row, column=2, value=f"CELLA INTERNA {i+1}")
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
                row += 1

                # Row 2: "MEDIA:"
                ws.cell(row=row, column=3, value="MEDIA:")
                ws.cell(row=row, column=4, value=f"{media_cella if media_cella is not None else 'N/A'}")
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                ws.cell(row=row, column=3).fill = media_fill  # Apply fill color to media description cell
                row += 1

                # Row 3: "MAX-MIN:", "massimo:", maximum value, "minimo:", minimum value
                ws.cell(row=row, column=3, value="MAX-MIN:")
                ws.cell(row=row, column=4, value="massimo:")
                ws.cell(row=row, column=5, value=f"{max_cella if max_cella is not None else 'N/A'}")
                ws.cell(row=row, column=6, value="minimo:")
                ws.cell(row=row, column=7, value=f"{min_cella if min_cella is not None else 'N/A'}")
                ws.cell(row=row, column=4).fill = max_fill  # Apply fill color to massimo description cell
                ws.cell(row=row, column=6).fill = min_fill  # Apply fill color to minimo description cell
                row += 1

            # Add two blank rows between balances
            row += 2

        # Adjust column widths for readability
        for col_num, col in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value)) for cell in col)
            ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

        # Save the workbook
        default_name = "report.xlsx"
        destination_file, _ = QFileDialog.getSaveFileName(
            self, 
            "Salva File Concatenato Come", 
            default_name, 
            "Log files (*.xlsx);;Tutti i file (*)",
        )
        if destination_file:
            # Ensure the file saved has the .xlsx extension
            if not destination_file.endswith('.xlsx'):
                destination_file += '.xlsx'

            wb.save(destination_file)
            l.log_file(999, f"Excel report saved as: {destination_file}")
            self.show_success_message(f"Report salvato come \"{destination_file}\"")
    
    def convert_to_dict(self, pesata, n):
        """Converte una lista in un dizionario con le chiavi appropriate."""
        return {
            'id_bilancia': n,
            'pesoTot': pesata[0],
            'pesoc1': pesata[1] if len(pesata) > 1 else '',
            'pesoc2': pesata[2] if len(pesata) > 2 else '',
            'pesoc3': pesata[3] if len(pesata) > 3 else '',
            'pesoc4': pesata[4] if len(pesata) > 4 else '',
        }   
        
    def db_page(self):
        
        home_widget = QWidget()
        home_layout = QHBoxLayout()
        home_widget.setLayout(home_layout)
        
        p1 = pi.ClickableWidget(f.get_img("trash.png"),     "Elimina dati database") 
        p1.clicked.connect(lambda: self.svuota_file(f.get_db()))
        p2 = pi.ClickableWidget(f.get_img("share.png"), "Esporta dati database") 
        p2.clicked.connect(self.show_save_db)
        
        home_layout.addStretch()
        home_layout.addWidget(p1)
        home_layout.addSpacing(50)
        home_layout.addWidget(p2)
        home_layout.addStretch()
        
        self.stacked_widget.addWidget(home_widget)
        
    def log_page(self):
        
        home_widget = QWidget()
        home_layout = QHBoxLayout()
        home_widget.setLayout(home_layout)
        
        p1 = pi.ClickableWidget(f.get_img("trash.png"), "Elimina DataLog") 
        p1.clicked.connect(self.svuota_log)
        p2 = pi.ClickableWidget(f.get_img("share.png"), "Esporta DataLog") 
        p2.clicked.connect(self.show_save_log)
        
        home_layout.addStretch()
        home_layout.addWidget(p1)
        home_layout.addSpacing(50)
        home_layout.addWidget(p2)
        home_layout.addStretch()
        
        self.stacked_widget.addWidget(home_widget)
         
    
    def sett_page(self):
        
        home_widget = QWidget()
        home_layout = QVBoxLayout()
        home_widget.setLayout(home_layout)
        home_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        h1 = QHBoxLayout()
        h2 = QHBoxLayout()
        h3 = QHBoxLayout()
        
        lab = QLabel("Valore massimo di controllo: ")
        lab.setObjectName("desc1")  
        
        self.kg = QSpinBox(self)
        self.kg.setRange(1, 400)
        val = self.master.master.binario.get_lettura_val()
        self.kg.setValue(val)
        self.kg.setSingleStep(1)
        self.kg.setMinimumWidth(100)
        
        kglab = QLabel("Kg")
        kglab.setObjectName("desc1")
        
        h1.addStretch()
        h1.addWidget(lab)
        h1.addSpacing(20)
        h1.addWidget(self.kg)
        h1.addWidget(kglab)
        h1.addStretch()
        
        self.controllo = QCheckBox("Controllo Interno Abilitato")
        self.controllo.setChecked(self.master.master.binario.get_lettura())
        h2.addStretch()
        h2.addWidget(self.controllo)
        h2.addStretch()
        
        self.tara_auto = QCheckBox("Auto Tara all'accensione")
        self.tara_auto.setChecked(self.master.master.binario.get_autotare())
        h3.addStretch()
        h3.addWidget(self.tara_auto)
        h3.addStretch()
        
        self.pulsante_salva = QPushButton("SALVA")
        self.pulsante_salva.setObjectName("bigd")
        self.pulsante_salva.clicked.connect(self.salva_bin)
        
        home_layout.addStretch()
        home_layout.addLayout(h1)
        home_layout.addStretch()
        home_layout.addLayout(h2)
        home_layout.addStretch()
        home_layout.addLayout(h3)
        home_layout.addStretch()
        home_layout.addWidget(self.pulsante_salva)
        home_layout.addStretch()
        
        self.stacked_widget.addWidget(home_widget)
        
    def salva_bin(self):
        self.master.master.binario.set_autoTara(self.tara_auto.isChecked())
        self.master.master.binario.set_lettura(self.controllo.isChecked(), self.kg.value())
            
    def clearLayout(self, layout):
        if layout is not None:
            while layout.count():
                item = layout.takeAt(0)
                widget = item.widget()
                if widget is not None:
                    widget.deleteLater()
                else:
                    self.clearLayout(item.layout())
                    
    def dataLog_operatione(self):
        self.stacked_widget.setCurrentIndex(2)
        self.master.contro_label(7)
    
    
    def db_operatione(self):
        self.stacked_widget.setCurrentIndex(3)
        self.master.contro_label(6)
    
    def svuota_log(self):
        self.svuota_file(f.get_app_log())
        self.svuota_file(f.get_thread_log())
        
    
    def show_save_db(self):
        
        # Prima, scegli il file da copiare
        db = f.get_db()
        
        default_name = "database.db"
        destination_file, _ = QFileDialog.getSaveFileName(self, 
                                                         "Salva File Concatenato Come", 
                                                         default_name, 
                                                         "Log files (*.db);;Tutti i file (*)",
                                                         )
        if destination_file:
            # Assicurarsi che il file salvato abbia l'estensione .log
            if not destination_file.endswith('.db'):
                destination_file += '.db'
            
            try:      
                # Leggi il contenuto dei due file
                with open(db, 'r') as file1, open(destination_file, 'w') as outfile:
                    outfile.write(file1.read())
                      # Aggiunge una nuova linea tra i contenuti dei due file
                    
                
                l.log_file(17, f"{destination_file}")
                
                self.show_success_message(f"Salvataggio del file \"{destination_file}\" db termianta con successo")
            except Exception as e:
                self.show_error_message(f"Errore durante il salvataggio del file: {e}")
                
            
    def show_save_log(self):
        # Prima, scegli il file da copiare
        log1 = f.get_app_log()
        log2 = f.get_thread_log()

        default_name = "concatenated_log.log"
        destination_file, _ = QFileDialog.getSaveFileName(self, 
                                                         "Salva File Concatenato Come", 
                                                         default_name, 
                                                         "Log files (*.log);;Tutti i file (*)",
                                                         )
        if destination_file:
            # Assicurarsi che il file salvato abbia l'estensione .log
            if not destination_file.endswith('.log'):
                destination_file += '.log'

            try:
                # Check if the source files exist and are readable
                if not os.path.exists(log1):
                    raise FileNotFoundError(f"File not found: {log1}")
                if not os.path.exists(log2):
                    raise FileNotFoundError(f"File not found: {log2}")

                # Open files and write content
                with open(log1, 'r') as file1, open(log2, 'r') as file2, open(destination_file, 'w') as outfile:
                    outfile.write(file1.read())
                    outfile.write("\n")  # Aggiunge una nuova linea tra i contenuti dei due file
                    outfile.write(file2.read())

                l.log_file(15, f"{destination_file}")
                self.show_success_message(f"Concatenazione e salvataggio del file \"{destination_file}\" log termianta con successo")
            except Exception as e:
                # Log the error and show an error message
                l.log_file(418, f"Errore durante la concatenazione dei file: {e}")
                self.show_error_message(f"Errore durante la concatenazione dei file: {e}")

    
    def svuota_file(self, percorso_file):
        try:
            # Apri il file in modalità di scrittura ('w') per svuotarlo
            with open(percorso_file, 'w') as file:
                pass  # Non scriviamo nulla, lasciamo semplicemente il file vuoto
            if percorso_file == f.get_db():
                self.master.master.ripristino_db()
                l.log_file(18)
            else:
                l.log_file(16)
            self.show_success_message(f"Il contenuto del file '{percorso_file}' è stato eliminato.")
        except Exception as e:
            self.show_error_message("Si è verificato un errore durante l'eliminazione del contenuto del file: {e}")
    
    def show_success_message(self, messaggio):
        ico = QIcon(f.get_img("close.png"))
        myModal = m.QCustomModals.SuccessModal(
            title="salvataggio eseguito",  # Title of the modal dialog
            parent=self,  # Parent widget to which the modal belongs
            position='top-right',  # Position to display the modal dialog
            closeIcon=ico,  # Path to the close icon image
            description=messaggio,  # Description text displayed in the modal dialog
            isClosable=False,  # Whether the modal dialog is closable (True or False)
            duration=3000  # Duration (in milliseconds) for which the modal dialog remains visible
        )

        # Show the modal
        myModal.show()
        
    def show_error_message(self, messaggio):
        ico = QIcon(f.get_img("close.png"))
        myModal = m.QCustomModals.ErrorModal(
            title="salvataggio eseguito",  # Title of the modal dialog
            parent=self,  # Parent widget to which the modal belongs
            position='top-right',  # Position to display the modal dialog
            closeIcon=ico,  # Path to the close icon image
            description=messaggio,  # Description text displayed in the modal dialog
            isClosable=False,  # Whether the modal dialog is closable (True or False)
            duration=3000  # Duration (in milliseconds) for which the modal dialog remains visible
        )

        # Show the modal
        myModal.show()